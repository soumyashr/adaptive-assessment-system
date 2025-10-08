#!/usr/bin/env python3
"""
Excel Template Generator for Question Banks
Creates formatted Excel templates with sample data and instructions

File: backend/scripts/xlsx_templates.py
"""

import pandas as pd
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from datetime import datetime


class TemplateGenerator:
    """Generate Excel templates for question upload"""

    def create_xlsx_template(self) -> io.BytesIO:
        """
        Create a comprehensive Excel template with multiple sheets

        Returns:
            BytesIO buffer containing the Excel file
        """
        buffer = io.BytesIO()

        # Create workbook and sheets
        wb = Workbook()

        # Remove default sheet and create our sheets
        wb.remove(wb.active)

        # Create sheets
        data_sheet = wb.create_sheet("Questions", 0)
        instructions_sheet = wb.create_sheet("Instructions", 1)
        examples_sheet = wb.create_sheet("Examples", 2)

        # Populate each sheet
        self._create_data_sheet(data_sheet)
        self._create_instructions_sheet(instructions_sheet)
        self._create_examples_sheet(examples_sheet)

        # Save to buffer
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def _create_data_sheet(self, ws):
        """Create the main data entry sheet with validation"""

        # Headers with ALL columns (including optional ones)
        headers = [
            'subject', 'question_id', 'question',
            'option_a', 'option_b', 'option_c', 'option_d',
            'answer', 'topic', 'tier',
            'discrimination_a', 'difficulty_b', 'guessing_c'
        ]

        # Column descriptions for comments
        descriptions = {
            'subject': 'Subject area (default: maths)',
            'question_id': 'Unique ID (auto-generated if blank)',
            'question': 'The question text (required)',
            'option_a': 'First answer option (required)',
            'option_b': 'Second answer option (required)',
            'option_c': 'Third answer option (required)',
            'option_d': 'Fourth answer option (required)',
            'answer': 'Correct answer: a, b, c, or d (required)',
            'topic': 'Topic or category (required)',
            'tier': 'Difficulty: C1(Easy), C2(Medium), C3(Hard), C4(Expert)',
            'discrimination_a': 'IRT discrimination (default: 1.5)',
            'difficulty_b': 'IRT difficulty (auto-set by tier)',
            'guessing_c': 'IRT guessing parameter (default: 0.25)'
        }

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)

            # Add comment with description
            if header in descriptions:
                cell.comment = Comment(descriptions[header], "Template Generator")
                cell.comment.width = 200
                cell.comment.height = 50

        # Apply header formatting
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add sample data rows with mathematical symbols
        sample_data = [
            ['maths', '', 'What is 2² + 3²?', '13', '14', '15', '16', 'a', 'Algebra', 'C1', '', '', ''],
            ['maths', '', 'Solve for x: x² = 16', 'x = 4', 'x = ±4', 'x = -4', 'x = 8', 'b', 'Equations', 'C2', '', '',
             ''],
            ['maths', '', 'What is the value of π (pi)?', '3.14', '3.141', '3.1416', '3.14159', 'd', 'Constants', 'C2',
             '', '', ''],
            ['maths', '', 'Calculate: √(25) × √(4)', '10', '20', '7', '100', 'a', 'Roots', 'C1', '', '', ''],
            ['maths', '', 'What is ∫x dx?', 'x²/2 + C', 'x² + C', '2x + C', 'x + C', 'a', 'Calculus', 'C3', '', '', ''],
        ]

        for row_idx, row_data in enumerate(sample_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Add data validations

        # Answer validation (must be a, b, c, or d)
        answer_validation = DataValidation(
            type="list",
            formula1='"a,b,c,d"',
            allow_blank=False,
            showDropDown=True,
            errorTitle='Invalid Answer',
            error='Answer must be a, b, c, or d'
        )
        answer_col = headers.index('answer') + 1
        answer_validation.add(f'{get_column_letter(answer_col)}2:{get_column_letter(answer_col)}1000')
        ws.add_data_validation(answer_validation)

        # Tier validation
        tier_validation = DataValidation(
            type="list",
            formula1='"C1,C2,C3,C4"',
            allow_blank=False,
            showDropDown=True,
            errorTitle='Invalid Tier',
            error='Tier must be C1, C2, C3, or C4'
        )
        tier_col = headers.index('tier') + 1
        tier_validation.add(f'{get_column_letter(tier_col)}2:{get_column_letter(tier_col)}1000')
        ws.add_data_validation(tier_validation)

        # IRT parameter validations
        if 'discrimination_a' in headers:
            disc_validation = DataValidation(
                type="decimal",
                operator="greaterThan",
                formula1='0',
                errorTitle='Invalid Discrimination',
                error='Discrimination must be positive'
            )
            disc_col = headers.index('discrimination_a') + 1
            disc_validation.add(f'{get_column_letter(disc_col)}2:{get_column_letter(disc_col)}1000')
            ws.add_data_validation(disc_validation)

        # Auto-adjust column widths
        column_widths = {
            'subject': 10,
            'question_id': 12,
            'question': 40,
            'option_a': 15,
            'option_b': 15,
            'option_c': 15,
            'option_d': 15,
            'answer': 8,
            'topic': 15,
            'tier': 6,
            'discrimination_a': 15,
            'difficulty_b': 12,
            'guessing_c': 10
        }

        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = column_widths.get(header, 12)

        # Apply borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=len(sample_data) + 1,
                                min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        # Color coding for required vs optional columns
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)

            # Required columns get darker blue
            required = ['question', 'option_a', 'option_b', 'option_c', 'option_d',
                        'answer', 'topic', 'tier']

            if header in required:
                cell.fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
            else:
                cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')

        # Freeze top row
        ws.freeze_panes = 'A2'

    def _create_instructions_sheet(self, ws):
        """Create instructions sheet"""

        instructions = [
            ["Question Bank Template Instructions", ""],
            ["", ""],
            ["OVERVIEW", ""],
            ["This template helps you create properly formatted question banks for the IRT assessment system.", ""],
            ["", ""],
            ["REQUIRED COLUMNS", ""],
            ["Column", "Description"],
            ["question", "The question text"],
            ["option_a", "First answer choice"],
            ["option_b", "Second answer choice"],
            ["option_c", "Third answer choice"],
            ["option_d", "Fourth answer choice"],
            ["answer", "Correct answer (must be: a, b, c, or d)"],
            ["topic", "Topic or category of the question"],
            ["tier", "Difficulty level: C1 (Easy), C2 (Medium), C3 (Hard), C4 (Expert)"],
            ["", ""],
            ["OPTIONAL COLUMNS", ""],
            ["Column", "Description"],
            ["subject", "Subject area (defaults to 'maths' if not provided)"],
            ["question_id", "Unique identifier (auto-generated if not provided)"],
            ["discrimination_a", "IRT discrimination parameter (defaults to 1.5)"],
            ["difficulty_b", "IRT difficulty parameter (auto-set based on tier)"],
            ["guessing_c", "IRT guessing parameter (defaults to 0.25 for 4-option MCQ)"],
            ["", ""],
            ["TIER DIFFICULTY MAPPING", ""],
            ["Tier", "Difficulty Value"],
            ["C1", "-1.0 (Easy)"],
            ["C2", "-0.5 (Medium)"],
            ["C3", "0.5 (Hard)"],
            ["C4", "1.0 (Expert)"],
            ["", ""],
            ["MATHEMATICAL SYMBOLS", ""],
            ["Excel preserves all mathematical symbols perfectly:", ""],
            ["• Superscripts: x², x³, 2⁴", ""],
            ["• Greek letters: π, θ, α, β, γ, Σ", ""],
            ["• Operators: √, ∞, ≤, ≥, ≠, ±", ""],
            ["• Fractions: ½, ⅓, ¼, ⅔, ¾", ""],
            ["", ""],
            ["TIPS", ""],
            ["1. Keep questions clear and concise", ""],
            ["2. Ensure all options are plausible", ""],
            ["3. Avoid 'All of the above' or 'None of the above' options", ""],
            ["4. Use consistent formatting across questions", ""],
            ["5. Group questions by topic for better organization", ""],
            ["", ""],
            ["DATA VALIDATION", ""],
            ["The template includes automatic validation for:", ""],
            ["• Answer column: Only accepts a, b, c, or d", ""],
            ["• Tier column: Only accepts C1, C2, C3, or C4", ""],
            ["• Discrimination: Must be positive", ""],
            ["", ""],
            ["UPLOADING", ""],
            ["1. Fill in your questions in the 'Questions' sheet", ""],
            ["2. Delete the sample rows (keep the header row)", ""],
            ["3. Save the file", ""],
            ["4. Upload through the admin interface", ""],
        ]

        # Write instructions
        for row_idx, row_data in enumerate(instructions, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Format headers
                if value in ["OVERVIEW", "REQUIRED COLUMNS", "OPTIONAL COLUMNS",
                             "TIER DIFFICULTY MAPPING", "MATHEMATICAL SYMBOLS",
                             "TIPS", "DATA VALIDATION", "UPLOADING"]:
                    cell.font = Font(bold=True, size=12, color='1F4788')
                elif row_idx == 1:  # Main title
                    cell.font = Font(bold=True, size=16, color='1F4788')
                elif value == "Column" or value == "Description" or value == "Tier" or value == "Difficulty Value":
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

        # Adjust column widths
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 50

    def _create_examples_sheet(self, ws):
        """Create sheet with example questions across different subjects"""

        # Headers
        headers = ['subject', 'question', 'option_a', 'option_b', 'option_c', 'option_d',
                   'answer', 'topic', 'tier']

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')

        # Example questions with various mathematical content
        examples = [
            # Mathematics
            ['maths', 'What is the derivative of x³?', '3x²', '2x³', 'x²', '3x', 'a', 'Calculus', 'C2'],
            ['maths', 'Solve: log₂(8) = ?', '2', '3', '4', '8', 'b', 'Logarithms', 'C2'],
            ['maths', 'What is sin(90°)?', '0', '1', '-1', '0.5', 'b', 'Trigonometry', 'C1'],
            ['maths', 'If f(x) = 2x + 3, what is f⁻¹(x)?', '(x-3)/2', '(x+3)/2', '2x-3', '3x+2', 'a', 'Functions',
             'C3'],

            # Physics
            ['physics', 'F = ma represents which law?', "Newton's 1st", "Newton's 2nd", "Newton's 3rd", "Hooke's Law",
             'b', 'Mechanics', 'C1'],
            ['physics', 'What is the speed of light in vacuum?', '3×10⁸ m/s', '3×10⁶ m/s', '3×10¹⁰ m/s', '3×10⁷ m/s',
             'a', 'Optics', 'C1'],

            # Chemistry
            ['chemistry', 'What is the atomic number of Carbon?', '6', '12', '14', '8', 'a', 'Periodic Table', 'C1'],
            ['chemistry', 'pH of a neutral solution at 25°C is:', '0', '7', '14', '1', 'b', 'Acids & Bases', 'C1'],

            # Complex Math
            ['maths', '∫(sin x)dx = ?', '-cos x + C', 'cos x + C', 'sin x + C', '-sin x + C', 'a', 'Integration', 'C2'],
            ['maths', 'lim(x→∞) (1 + 1/x)ˣ = ?', 'e', '1', '∞', '0', 'a', 'Limits', 'C3'],
            ['maths', 'The sum of angles in a triangle is:', '90°', '180°', '360°', '270°', 'b', 'Geometry', 'C1'],
            ['maths', 'i² = ? (where i is imaginary unit)', '1', '-1', '0', 'i', 'b', 'Complex Numbers', 'C2'],
        ]

        # Write examples
        for row_idx, row_data in enumerate(examples, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Apply borders and adjust widths
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=len(examples) + 1,
                                min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        # Column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 6

        # Color code tiers
        tier_colors = {
            'C1': 'C6EFCE',  # Light green
            'C2': 'FFEB9C',  # Light yellow
            'C3': 'FFC7CE',  # Light red
            'C4': 'D9B3FF'  # Light purple
        }

        tier_col = headers.index('tier') + 1
        for row_idx in range(2, len(examples) + 2):
            cell = ws.cell(row=row_idx, column=tier_col)
            tier_value = cell.value
            if tier_value in tier_colors:
                cell.fill = PatternFill(
                    start_color=tier_colors[tier_value],
                    end_color=tier_colors[tier_value],
                    fill_type='solid'
                )

        # Freeze header row
        ws.freeze_panes = 'A2'

    def create_csv_template(self) -> str:
        """
        Create a simple CSV template

        Returns:
            CSV content as string
        """
        # Create dataframe with sample data
        data = {
            'subject': ['maths'] * 5,
            'question_id': ['', '', '', '', ''],
            'question': [
                'What is 2 + 2?',
                'What is 5 × 3?',
                'What is 10 - 4?',
                'What is 12 ÷ 3?',
                'What is 3²?'
            ],
            'option_a': ['3', '10', '5', '3', '6'],
            'option_b': ['4', '15', '6', '4', '9'],
            'option_c': ['5', '20', '7', '5', '12'],
            'option_d': ['6', '25', '8', '6', '15'],
            'answer': ['b', 'b', 'b', 'b', 'b'],
            'topic': ['Addition', 'Multiplication', 'Subtraction', 'Division', 'Powers'],
            'tier': ['C1', 'C1', 'C1', 'C1', 'C2'],
            'discrimination_a': ['', '', '', '', ''],
            'difficulty_b': ['', '', '', '', ''],
            'guessing_c': ['', '', '', '', '']
        }

        df = pd.DataFrame(data)

        # Convert to CSV
        return df.to_csv(index=False)


def generate_template_file(output_path: str = None, format: str = 'xlsx'):
    """
    Generate a template file

    Args:
        output_path: Path to save the file (optional)
        format: 'xlsx' or 'csv'
    """
    generator = TemplateGenerator()

    if format == 'xlsx':
        buffer = generator.create_xlsx_template()

        if output_path:
            with open(output_path, 'wb') as f:
                f.write(buffer.getvalue())
        else:
            # Default path
            with open('question_template.xlsx', 'wb') as f:
                f.write(buffer.getvalue())

        print(f"✅ Excel template created: {output_path or 'question_template.xlsx'}")

    elif format == 'csv':
        content = generator.create_csv_template()

        if output_path:
            with open(output_path, 'w', encoding='utf-8-sig') as f:
                f.write(content)
        else:
            with open('question_template.csv', 'w', encoding='utf-8-sig') as f:
                f.write(content)

        print(f"✅ CSV template created: {output_path or 'question_template.csv'}")


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='Generate question bank templates')
    parser.add_argument('--format', choices=['xlsx', 'csv'], default='xlsx',
                        help='Template format (default: xlsx)')
    parser.add_argument('--output', '-o', help='Output file path')

    args = parser.parse_args()

    generate_template_file(args.output, args.format)