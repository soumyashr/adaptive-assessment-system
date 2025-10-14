import io
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class TemplateGenerator:
    """Generate Excel templates for item bank uploads"""

    @staticmethod
    def create_item_bank_template() -> io.BytesIO:
        """
        Create an Excel template for item bank upload

        Returns:
            BytesIO: Excel file in memory
        """
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Item Bank Template"

        # Define headers
        headers = [
            "id",
            "content",
            "option_1",
            "option_2",
            "option_3",
            "option_4",
            "correct_answer",
            "difficulty",
            "discrimination"
        ]

        # Style for headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Add sample data row
        sample_data = [
            "q1",
            "What is 2 + 2?",
            "2",
            "3",
            "4",
            "5",
            "3",
            "0.3",
            "1.0"
        ]

        for col_num, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(vertical="center")

        # Add instructions in a separate sheet
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ["Field", "Description", "Required", "Example"],
            ["id", "Unique identifier for the question", "Yes", "q1, q2, math_001"],
            ["content", "The question text", "Yes", "What is 2 + 2?"],
            ["option_1", "First answer option", "Yes", "2"],
            ["option_2", "Second answer option", "Yes", "3"],
            ["option_3", "Third answer option", "Yes", "4"],
            ["option_4", "Fourth answer option", "Yes", "5"],
            ["correct_answer", "Index of correct answer (1-4)", "Yes", "3 (means option_3 is correct)"],
            ["difficulty", "Question difficulty (0.0 to 1.0)", "Yes", "0.3 (easier) to 0.9 (harder)"],
            ["discrimination", "Discrimination parameter (0.5 to 2.0)", "Yes", "1.0 (typical value)"],
            ["", "", "", ""],
            ["Notes:", "", "", ""],
            ["- difficulty: Lower values = easier questions (0.1-0.3), Higher values = harder questions (0.7-0.9)", "",
             "", ""],
            ["- discrimination: Typically use 1.0. Higher values (1.5-2.0) = better at differentiating ability levels",
             "", "", ""],
            ["- correct_answer: Use 1 for option_1, 2 for option_2, 3 for option_3, 4 for option_4", "", "", ""],
        ]

        for row_num, row_data in enumerate(instructions, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_instructions.cell(row=row_num, column=col_num)
                cell.value = value
                if row_num == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                cell.border = border

        # Adjust column widths
        for ws_temp in [ws, ws_instructions]:
            for column in ws_temp.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_temp.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    @staticmethod
    def create_csv_template() -> str:
        """
        Create a CSV template for item bank upload

        Returns:
            str: CSV content
        """
        csv_content = """id,content,option_1,option_2,option_3,option_4,correct_answer,difficulty,discrimination
q1,What is 2 + 2?,2,3,4,5,3,0.3,1.0
q2,What is the capital of France?,London,Berlin,Paris,Madrid,3,0.4,1.2
q3,What is 10 x 5?,25,50,75,100,2,0.5,1.0"""
        return csv_content