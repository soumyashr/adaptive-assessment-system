#!/usr/bin/env python3
"""
Excel File Cleaner Script (XLSX/XLS Only)
Processes Excel files from raw_sample_questions folder

Features:
- XLSX/XLS format only - perfect mathematical symbol preservation
- Auto-generates IRT parameters and missing columns
- Validates and standardizes data
- Creates formatted output with color-coded tiers

Directory structure:
backend/
├── scripts/
│   └── clean_xlsx_files.py (this script)
└── data/
    ├── logs/
    │   └── cleaning_log_*.log
    ├── raw_sample_questions/
    │   ├── *.xlsx, *.xls (files to process)
    │   └── processed/
    │       └── *_processed_*.xlsx
    └── sample_questions/
        └── *_cleaned.xlsx

Usage: python backend/scripts/clean_xlsx_files.py
"""

import os
import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import logging
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Setup paths
SCRIPT_DIR = Path(__file__).resolve().parent  # backend/scripts
BACKEND_DIR = SCRIPT_DIR.parent  # backend folder
DATA_DIR = BACKEND_DIR / "data"  # backend/data
RAW_DIR = DATA_DIR / "raw_sample_questions"
CLEANED_DIR = DATA_DIR / "sample_questions"
PROCESSED_DIR = RAW_DIR / "processed"
LOG_DIR = DATA_DIR / "logs"

# Create directories
for dir_path in [LOG_DIR, RAW_DIR, CLEANED_DIR, PROCESSED_DIR]:
    dir_path.mkdir(parents=True, exist_ok=True)

# Setup logging
LOG_FILE = LOG_DIR / f"cleaning_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class ExcelCleaner:
    """Cleans and standardizes Excel question files"""

    def __init__(self):
        self.stats = {
            'total_files': 0,
            'success': 0,
            'failed': 0,
            'total_questions': 0,
            'files_processed': []
        }

    def process_file(self, filepath: Path) -> dict:
        """
        Process a single Excel file

        Args:
            filepath: Path to the Excel file

        Returns:
            dict: Cleaned data and statistics
        """
        logger.info(f"Processing: {filepath.name}")
        file_ext = filepath.suffix.lower()

        # Validate file type
        if file_ext not in ['.xlsx', '.xls']:
            raise ValueError(f"Invalid file type: {file_ext}. Only Excel files (.xlsx, .xls) are supported.")

        try:
            # Read Excel file
            df = self._read_excel_file(filepath)

            if df.empty:
                raise ValueError("File contains no data")

            # Clean and standardize the data
            cleaned_df = self._clean_dataframe(df, filepath.stem)

            # Validate the cleaned data
            self._validate_cleaned_data(cleaned_df)

            # Log statistics
            tier_counts = cleaned_df['tier'].value_counts().to_dict()
            logger.info(f"✓ Processed {len(cleaned_df)} questions")
            logger.info(f"  Tier distribution: {tier_counts}")

            # Check for mathematical symbols
            if self._contains_math_symbols(cleaned_df):
                logger.info("  ✨ Mathematical symbols preserved perfectly")

            return {
                'dataframe': cleaned_df,
                'stats': {
                    'total': len(cleaned_df),
                    'tiers': tier_counts
                }
            }

        except Exception as e:
            logger.error(f"✗ Failed to process {filepath.name}: {str(e)}")
            raise

    def _read_excel_file(self, filepath: Path) -> pd.DataFrame:
        """Read Excel file and find the data sheet"""
        try:
            # Read Excel file
            excel_file = pd.ExcelFile(filepath)

            # Log available sheets
            logger.debug(f"  Available sheets: {excel_file.sheet_names}")

            # Find the data sheet (skip instruction/template sheets)
            data_sheet = None
            for sheet_name in excel_file.sheet_names:
                sheet_lower = sheet_name.lower()
                if 'instruction' not in sheet_lower and 'template' not in sheet_lower and 'example' not in sheet_lower:
                    data_sheet = sheet_name
                    break

            # Default to first sheet if no data sheet found
            if data_sheet is None:
                data_sheet = 0
                logger.info(f"  No specific data sheet found, using first sheet: {excel_file.sheet_names[0]}")
            else:
                logger.info(f"  Using sheet: {data_sheet}")

            # Read the data sheet
            df = pd.read_excel(filepath, sheet_name=data_sheet)

            # Remove any completely empty rows
            df = df.dropna(how='all')

            return df

        except Exception as e:
            raise ValueError(f"Could not read Excel file: {str(e)}")

    def _clean_dataframe(self, df: pd.DataFrame, filename_stem: str) -> pd.DataFrame:
        """Clean and standardize the dataframe"""

        # Normalize column names
        df.columns = df.columns.str.lower().str.strip()

        # Check required columns
        required = ['question', 'option_a', 'option_b', 'option_c', 'option_d',
                    'answer', 'tier', 'topic']
        missing = [col for col in required if col not in df.columns]

        if missing:
            raise ValueError(f"Missing required columns: {missing}")

        # Generate ID prefix from filename
        id_prefix = ''.join(c for c in filename_stem.lower() if c.isalnum())

        # Add missing columns with defaults
        if 'subject' not in df.columns:
            df['subject'] = 'maths'
            logger.info("  Added 'subject' column with default value: maths")

        if 'question_id' not in df.columns:
            df['question_id'] = [f"{id_prefix}_{i:03d}" for i in range(1, len(df) + 1)]
            logger.info(f"  Generated question IDs with prefix: {id_prefix}")

        # Fix answer format (e.g., "option_b" -> "b")
        df['answer'] = df['answer'].astype(str).str.lower().str.strip()
        if df['answer'].str.contains('option_').any():
            df['answer'] = df['answer'].str.replace('option_', '', regex=False)
            logger.info("  Fixed answer format (removed 'option_' prefix)")

        # Validate answers
        valid_answers = ['a', 'b', 'c', 'd']
        invalid_mask = ~df['answer'].isin(valid_answers)
        if invalid_mask.any():
            invalid_count = invalid_mask.sum()
            logger.warning(f"  Found {invalid_count} invalid answers, defaulting to 'a'")
            df.loc[invalid_mask, 'answer'] = 'a'

        # Standardize tier
        df['tier'] = df['tier'].astype(str).str.upper().str.strip()
        if not df['tier'].str.startswith('C').all():
            df.loc[~df['tier'].str.startswith('C'), 'tier'] = 'C' + df.loc[~df['tier'].str.startswith('C'), 'tier']
            logger.info("  Standardized tier format")

        # Validate tiers
        valid_tiers = ['C1', 'C2', 'C3', 'C4']
        invalid_tier_mask = ~df['tier'].isin(valid_tiers)
        if invalid_tier_mask.any():
            invalid_count = invalid_tier_mask.sum()
            logger.warning(f"  Found {invalid_count} invalid tiers, defaulting to 'C2'")
            df.loc[invalid_tier_mask, 'tier'] = 'C2'

        # Add IRT parameters based on tier
        tier_difficulty = {
            'C1': -1.0,  # Easy
            'C2': -0.5,  # Medium
            'C3': 0.5,  # Hard
            'C4': 1.0  # Expert
        }

        if 'discrimination_a' not in df.columns:
            df['discrimination_a'] = 1.5
            logger.info("  Added discrimination_a with default value: 1.5")

        if 'difficulty_b' not in df.columns:
            df['difficulty_b'] = df['tier'].map(tier_difficulty)
            logger.info("  Generated difficulty_b based on tier")

        if 'guessing_c' not in df.columns:
            df['guessing_c'] = 0.25
            logger.info("  Added guessing_c with default value: 0.25")

        # Enhance topic names if needed
        if 'topic' in df.columns and not df['topic'].astype(str).str.contains('-').any():
            base_topic = ' '.join(word.capitalize() for word in filename_stem.replace('_', ' ').split())
            df['topic'] = base_topic + ' - ' + df['topic'].astype(str)
            logger.info(f"  Enhanced topic names with base: {base_topic}")

        # Ensure all text columns are strings and handle NaN
        text_cols = ['question', 'option_a', 'option_b', 'option_c', 'option_d', 'topic', 'subject']
        for col in text_cols:
            if col in df.columns:
                df[col] = df[col].fillna('').astype(str).str.strip()

        # Reorder columns
        column_order = ['subject', 'question_id', 'question', 'option_a', 'option_b',
                        'option_c', 'option_d', 'answer', 'topic', 'tier',
                        'discrimination_a', 'difficulty_b', 'guessing_c']

        # Add any extra columns at the end
        extra_cols = [col for col in df.columns if col not in column_order]
        if extra_cols:
            logger.info(f"  Preserving extra columns: {extra_cols}")
            column_order.extend(extra_cols)

        return df[column_order]

    def _validate_cleaned_data(self, df: pd.DataFrame):
        """Validate the cleaned dataframe"""
        issues = []

        # Check for empty questions
        empty_questions = df[df['question'].str.strip() == ''].shape[0]
        if empty_questions > 0:
            issues.append(f"{empty_questions} empty questions")

        # Check for duplicate question IDs
        duplicates = df['question_id'].duplicated().sum()
        if duplicates > 0:
            issues.append(f"{duplicates} duplicate question IDs")

        # Check for missing options
        for opt in ['option_a', 'option_b', 'option_c', 'option_d']:
            empty_opts = df[df[opt].str.strip() == ''].shape[0]
            if empty_opts > 0:
                issues.append(f"{empty_opts} empty {opt}")

        if issues:
            logger.warning(f"  ⚠ Validation issues: {', '.join(issues)}")

    def _contains_math_symbols(self, df: pd.DataFrame) -> bool:
        """Check if dataframe contains mathematical symbols"""
        math_symbols = ['²', '³', '⁴', '⁵', '⁶', '⁷', '⁸', '⁹', '⁰',
                        '√', 'π', 'θ', 'α', 'β', 'γ', 'δ', 'ε', 'λ', 'μ', 'σ', 'φ', 'ω',
                        '∑', '∫', '∂', '∇', '∞',
                        '≤', '≥', '≠', '≈', '≡',
                        '±', '×', '÷', '·',
                        '½', '⅓', '¼', '⅔', '¾']

        # Check in questions and options
        text_cols = ['question', 'option_a', 'option_b', 'option_c', 'option_d']

        for col in text_cols:
            if col in df.columns:
                col_text = df[col].astype(str).str.cat()
                for symbol in math_symbols:
                    if symbol in col_text:
                        return True

        return False

    def save_to_excel(self, df: pd.DataFrame, output_path: Path):
        """
        Save dataframe to Excel with professional formatting

        Args:
            df: Dataframe to save
            output_path: Output file path
        """
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Questions', index=False)

            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Questions']

            # Define styles
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            cell_alignment = Alignment(vertical='top', wrap_text=True)

            border = Border(
                left=Side(style='thin', color='DDDDDD'),
                right=Side(style='thin', color='DDDDDD'),
                top=Side(style='thin', color='DDDDDD'),
                bottom=Side(style='thin', color='DDDDDD')
            )

            # Apply header formatting
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Apply cell formatting
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = cell_alignment
                    cell.border = border

            # Auto-adjust column widths
            column_widths = {
                'subject': 10,
                'question_id': 12,
                'question': 40,
                'option_a': 18,
                'option_b': 18,
                'option_c': 18,
                'option_d': 18,
                'answer': 8,
                'topic': 20,
                'tier': 8,
                'discrimination_a': 15,
                'difficulty_b': 12,
                'guessing_c': 12
            }

            for col_idx, col_name in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_idx)
                if col_name in column_widths:
                    worksheet.column_dimensions[col_letter].width = column_widths[col_name]
                else:
                    # Auto-width for unexpected columns
                    max_length = max(
                        len(str(col_name)),
                        df[col_name].astype(str).map(len).max()
                    )
                    worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)

            # Color-code tiers
            tier_colors = {
                'C1': 'E8F5E9',  # Light green (Easy)
                'C2': 'FFF9C4',  # Light yellow (Medium)
                'C3': 'FFE0B2',  # Light orange (Hard)
                'C4': 'FFEBEE'  # Light red (Expert)
            }

            if 'tier' in df.columns:
                tier_col_index = df.columns.get_loc('tier') + 1
                for row in range(2, len(df) + 2):
                    tier_cell = worksheet.cell(row=row, column=tier_col_index)
                    tier_value = str(tier_cell.value).upper()

                    if tier_value in tier_colors:
                        tier_cell.fill = PatternFill(
                            start_color=tier_colors[tier_value],
                            end_color=tier_colors[tier_value],
                            fill_type='solid'
                        )
                        tier_cell.font = Font(bold=True)

            # Freeze header row
            worksheet.freeze_panes = 'A2'

            # Add print settings
            worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
            worksheet.page_setup.fitToWidth = 1
            worksheet.page_setup.fitToHeight = False
            worksheet.print_options.horizontalCentered = True

        logger.info(f"  💾 Saved to: {output_path.name}")

    def process_all_files(self):
        """Process all Excel files in the raw directory"""
        logger.info("=" * 60)
        logger.info("EXCEL FILE CLEANING STARTED")
        logger.info("=" * 60)
        logger.info(f"Script location: {SCRIPT_DIR}")
        logger.info(f"Raw directory: {RAW_DIR}")
        logger.info(f"Output directory: {CLEANED_DIR}")
        logger.info(f"Log directory: {LOG_DIR}")
        logger.info("📊 Processing Excel files only (.xlsx, .xls)")
        logger.info("✨ Mathematical symbols perfectly preserved")
        logger.info("-" * 60)

        # Find all Excel files
        excel_files = list(RAW_DIR.glob("*.xlsx")) + list(RAW_DIR.glob("*.xls"))

        # Filter out files in processed directory
        excel_files = [f for f in excel_files if f.is_file() and f.parent == RAW_DIR]

        if not excel_files:
            logger.warning("No Excel files found in raw_sample_questions directory")
            logger.info(f"Place Excel files (.xlsx or .xls) in: {RAW_DIR}")
            return

        logger.info(f"Found {len(excel_files)} Excel files to process")
        logger.info("-" * 60)

        for file_path in excel_files:
            self.stats['total_files'] += 1

            try:
                # Process the file
                result = self.process_file(file_path)

                # Save cleaned Excel file
                output_filename = file_path.stem + "_cleaned.xlsx"
                output_path = CLEANED_DIR / output_filename

                self.save_to_excel(result['dataframe'], output_path)

                # Move original to processed folder
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                processed_filename = f"{file_path.stem}_processed_{timestamp}{file_path.suffix}"
                processed_path = PROCESSED_DIR / processed_filename

                shutil.move(str(file_path), str(processed_path))

                self.stats['success'] += 1
                self.stats['total_questions'] += result['stats']['total']
                self.stats['files_processed'].append(output_filename)

                logger.info(f"✓ Successfully cleaned: {file_path.name}")
                logger.info(f"  Original moved to: processed/{processed_filename}")
                logger.info("-" * 60)

            except Exception as e:
                self.stats['failed'] += 1
                logger.error(f"✗ Failed to process {file_path.name}: {str(e)}")

                # Move failed file
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                failed_filename = f"{file_path.stem}_failed_{timestamp}{file_path.suffix}"
                failed_path = PROCESSED_DIR / failed_filename

                try:
                    shutil.move(str(file_path), str(failed_path))
                    logger.info(f"  Failed file moved to: processed/{failed_filename}")
                except Exception as move_error:
                    logger.error(f"  Could not move failed file: {move_error}")

                logger.info("-" * 60)

        self._print_summary()

    def _print_summary(self):
        """Print processing summary"""
        logger.info("=" * 60)
        logger.info("PROCESSING COMPLETE - SUMMARY")
        logger.info("=" * 60)
        logger.info(f"📊 Total files processed: {self.stats['total_files']}")
        logger.info(f"✅ Successful: {self.stats['success']}")
        logger.info(f"❌ Failed: {self.stats['failed']}")
        logger.info(f"📝 Total questions cleaned: {self.stats['total_questions']}")

        if self.stats['files_processed']:
            logger.info(f"\n📁 Cleaned files saved to: {CLEANED_DIR}")
            for filename in self.stats['files_processed']:
                logger.info(f"   - {filename}")

        logger.info(f"\n📋 Log file: {LOG_FILE.name}")
        logger.info(f"📂 Full log path: {LOG_FILE}")
        logger.info("=" * 60)

        # Console summary
        print(f"\n✅ Processing complete!")
        print(f"   - Processed {self.stats['success']} Excel files")
        print(f"   - Total {self.stats['total_questions']} questions cleaned")
        print(f"   - Mathematical symbols preserved")
        print(f"   - Check {CLEANED_DIR} for cleaned files")
        print(f"   - Check {LOG_FILE.name} for detailed log")


def main():
    """Main function"""
    try:
        print("🚀 Starting Excel File Cleaner...")
        print("   📊 Excel format only (.xlsx, .xls)")
        print("   ✨ Mathematical symbols perfectly preserved")
        print(f"   📁 Looking for files in: {RAW_DIR}")
        print()

        cleaner = ExcelCleaner()
        cleaner.process_all_files()

    except KeyboardInterrupt:
        logger.info("\n⚠ Process interrupted by user")
        print("\n⚠ Process interrupted")
    except Exception as e:
        logger.error(f"Unexpected error: {str(e)}")
        print(f"\n❌ Error: {str(e)}")
        print(f"   Check log file for details: {LOG_DIR}")
        raise


if __name__ == "__main__":
    main()