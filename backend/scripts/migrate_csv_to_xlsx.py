#!/usr/bin/env python3
"""
CSV to XLSX Migration Tool
Converts existing CSV files to Excel format with proper formatting

Features:
- Preserves all data and encoding
- Adds Excel formatting and styles
- Handles mathematical symbols
- Creates backup of original files
- Batch processing support

Usage:
    # Run from backend/ directory or backend/scripts/ directory
    python scripts/migrate_csv_to_xlsx.py --all

    # Or from backend/scripts/ directory
    python migrate_csv_to_xlsx.py --all

    # Migrate specific file
    python migrate_csv_to_xlsx.py path/to/questions.csv
"""

import pandas as pd
import numpy as np
from pathlib import Path
import argparse
import shutil
from datetime import datetime
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def find_project_root():
    """Find the project root directory (backend/)"""
    current = Path(__file__).resolve()

    # If script is in backend/scripts/, go up one level
    if current.parent.name == 'scripts':
        return current.parent.parent

    # If script is in backend/, use that
    if current.parent.name == 'backend':
        return current.parent

    # Otherwise, assume current directory
    return current.parent


class CSVToXLSXMigrator:
    """Migrate CSV files to XLSX format"""

    def __init__(self, create_backup=True):
        """
        Initialize migrator

        Args:
            create_backup: Whether to create backups of original files
        """
        self.create_backup = create_backup
        self.stats = {
            'total': 0,
            'success': 0,
            'failed': 0,
            'skipped': 0
        }

    def migrate_file(self, csv_path: Path, output_path: Path = None) -> bool:
        """
        Migrate a single CSV file to XLSX

        Args:
            csv_path: Path to CSV file
            output_path: Optional output path (defaults to same name with .xlsx)

        Returns:
            True if successful, False otherwise
        """
        if not csv_path.exists():
            logger.error(f"File not found: {csv_path}")
            return False

        if not csv_path.suffix.lower() == '.csv':
            logger.warning(f"Not a CSV file: {csv_path}")
            return False

        # Default output path
        if output_path is None:
            output_path = csv_path.with_suffix('.xlsx')

        logger.info(f"Migrating: {csv_path.name} → {output_path.name}")

        try:
            # Create backup if requested
            if self.create_backup:
                backup_path = csv_path.with_suffix('.csv.bak')
                if not backup_path.exists():
                    shutil.copy2(csv_path, backup_path)
                    logger.info(f"  Created backup: {backup_path.name}")

            # Read CSV with encoding detection
            df = self._read_csv_with_encoding(csv_path)

            # Clean and validate data
            df = self._clean_dataframe(df)

            # Check for mathematical symbols
            has_math = self._check_math_symbols(df)
            if has_math:
                logger.info("  ✨ Mathematical symbols detected and preserved")

            # Write to Excel with formatting
            self._write_formatted_excel(df, output_path)

            logger.info(f"  ✅ Successfully migrated: {output_path.name}")
            logger.info(f"     Rows: {len(df)}, Columns: {len(df.columns)}")

            return True

        except Exception as e:
            logger.error(f"  ❌ Migration failed: {str(e)}")
            import traceback
            logger.debug(traceback.format_exc())
            return False

    def _read_csv_with_encoding(self, csv_path: Path) -> pd.DataFrame:
        """Read CSV with automatic encoding detection"""
        encodings = ['utf-8-sig', 'utf-8', 'cp1252', 'iso-8859-1', 'latin-1']

        for encoding in encodings:
            try:
                df = pd.read_csv(csv_path, encoding=encoding)
                logger.debug(f"  Read with {encoding} encoding")

                # Clean any wrapper quotes
                for col in df.columns:
                    if df[col].dtype == 'object':
                        df[col] = df[col].str.strip('"').str.strip()

                return df

            except (UnicodeDecodeError, pd.errors.ParserError):
                continue

        # Last resort - read with error replacement
        logger.warning(f"  Using fallback encoding with error replacement")
        return pd.read_csv(csv_path, encoding='utf-8', errors='replace')

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and standardize dataframe"""

        # Normalize column names
        df.columns = df.columns.str.strip()

        # Handle common issues
        if 'answer' in df.columns:
            # Fix answer format (option_a -> a)
            df['answer'] = df['answer'].astype(str).str.lower().str.strip()
            df['answer'] = df['answer'].str.replace('option_', '', regex=False)

        if 'tier' in df.columns:
            # Standardize tier format
            df['tier'] = df['tier'].astype(str).str.upper().str.strip()
            # Add 'C' prefix if missing
            mask = ~df['tier'].str.match(r'^C\d+$', na=False)
            df.loc[mask & df['tier'].str.match(r'^\d+$', na=False), 'tier'] = 'C' + df.loc[
                mask & df['tier'].str.match(r'^\d+$', na=False), 'tier']

        # Remove any completely empty rows
        df = df.dropna(how='all')

        # Handle NaN values in text columns
        text_columns = df.select_dtypes(include=['object']).columns
        for col in text_columns:
            df[col] = df[col].fillna('')

        return df

    def _check_math_symbols(self, df: pd.DataFrame) -> bool:
        """Check if dataframe contains mathematical symbols"""
        math_symbols = ['²', '³', '⁴', '√', 'π', 'θ', 'α', 'β', 'γ', '∑', '∫',
                        '≤', '≥', '≠', '±', '∞', '×', '÷', '°']

        # Check text columns
        text_cols = df.select_dtypes(include=['object']).columns

        for col in text_cols:
            for symbol in math_symbols:
                if df[col].astype(str).str.contains(symbol, regex=False, na=False).any():
                    return True

        return False

    def _write_formatted_excel(self, df: pd.DataFrame, output_path: Path):
        """Write dataframe to Excel with formatting"""

        # Write to Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Data', index=False)

            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Data']

            # Apply formatting
            self._apply_excel_formatting(worksheet, df)

    def _apply_excel_formatting(self, worksheet, df: pd.DataFrame):
        """Apply formatting to Excel worksheet"""

        # Define styles
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        cell_alignment = Alignment(vertical='top', wrap_text=True)

        border_style = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Apply header formatting
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style

        # Apply cell formatting and borders
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = cell_alignment
                cell.border = border_style

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            # Set width (min 10, max 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Special formatting for tier column if present
        if 'tier' in df.columns:
            tier_col_idx = df.columns.get_loc('tier') + 1
            tier_colors = {
                'C1': 'C6EFCE',  # Light green
                'C2': 'FFEB9C',  # Light yellow
                'C3': 'FFC7CE',  # Light red
                'C4': 'D9B3FF'  # Light purple
            }

            for row_num in range(2, len(df) + 2):
                cell = worksheet.cell(row=row_num, column=tier_col_idx)
                tier_value = str(cell.value).upper() if cell.value else ''

                if tier_value in tier_colors:
                    cell.fill = PatternFill(
                        start_color=tier_colors[tier_value],
                        end_color=tier_colors[tier_value],
                        fill_type='solid'
                    )

        # Freeze top row
        worksheet.freeze_panes = 'A2'

        # Set print settings
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = False

    def migrate_directory(self, directory: Path, recursive: bool = False) -> dict:
        """
        Migrate all CSV files in a directory

        Args:
            directory: Directory path
            recursive: Whether to process subdirectories

        Returns:
            Migration statistics
        """
        if not directory.exists():
            logger.error(f"Directory not found: {directory}")
            logger.error(f"Looked in: {directory.absolute()}")
            return self.stats

        # Find CSV files
        if recursive:
            csv_files = list(directory.rglob("*.csv"))
        else:
            csv_files = list(directory.glob("*.csv"))

        # Filter out backup files
        csv_files = [f for f in csv_files if not f.name.endswith('.bak')]

        if not csv_files:
            logger.info(f"No CSV files found in {directory}")
            logger.info(f"Searched in: {directory.absolute()}")
            return self.stats

        logger.info(f"Found {len(csv_files)} CSV files to migrate")
        logger.info("=" * 50)

        for csv_file in csv_files:
            self.stats['total'] += 1

            # Skip if XLSX already exists and is newer
            xlsx_file = csv_file.with_suffix('.xlsx')
            if xlsx_file.exists():
                csv_mtime = csv_file.stat().st_mtime
                xlsx_mtime = xlsx_file.stat().st_mtime

                if xlsx_mtime > csv_mtime:
                    logger.info(f"Skipping {csv_file.name} (XLSX is newer)")
                    self.stats['skipped'] += 1
                    continue

            # Migrate the file
            if self.migrate_file(csv_file):
                self.stats['success'] += 1
            else:
                self.stats['failed'] += 1

            logger.info("-" * 50)

        return self.stats

    def print_summary(self):
        """Print migration summary"""
        print("\n" + "=" * 50)
        print("MIGRATION SUMMARY")
        print("=" * 50)
        print(f"Total files:     {self.stats['total']}")
        print(f"✅ Migrated:     {self.stats['success']}")
        print(f"⏭️  Skipped:      {self.stats['skipped']}")
        print(f"❌ Failed:       {self.stats['failed']}")

        if self.stats['success'] > 0:
            print("\n✨ Migration complete! Excel files preserve all mathematical symbols.")

        if self.create_backup:
            print("\n💾 Original CSV files backed up with .bak extension")


def main():
    parser = argparse.ArgumentParser(
        description='Migrate CSV files to Excel format (XLSX)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Migrate all CSVs in backend/data/sample_questions/
  python migrate_csv_to_xlsx.py --all

  # Migrate single file
  python migrate_csv_to_xlsx.py questions.csv

  # Migrate all CSVs in specific directory
  python migrate_csv_to_xlsx.py --directory ./data

  # Migrate recursively
  python migrate_csv_to_xlsx.py --directory ./data --recursive

  # Migrate without creating backups
  python migrate_csv_to_xlsx.py --no-backup --all

Note: Run this script from backend/scripts/ directory or backend/ directory
        """
    )

    parser.add_argument(
        'csv_file',
        nargs='?',
        help='Path to CSV file to migrate'
    )

    parser.add_argument(
        '--directory', '-d',
        help='Migrate all CSV files in directory'
    )

    parser.add_argument(
        '--recursive', '-r',
        action='store_true',
        help='Process subdirectories recursively'
    )

    parser.add_argument(
        '--no-backup',
        action='store_true',
        help='Do not create backup files'
    )

    parser.add_argument(
        '--all', '-a',
        action='store_true',
        help='Migrate all CSVs in backend/data/sample_questions/'
    )

    args = parser.parse_args()

    # Validate arguments
    if not any([args.csv_file, args.directory, args.all]):
        parser.print_help()
        sys.exit(1)

    # Create migrator
    migrator = CSVToXLSXMigrator(create_backup=not args.no_backup)

    # Perform migration
    if args.all:
        # Find project root and construct path to sample_questions
        project_root = find_project_root()
        data_dir = project_root / 'data' / 'sample_questions'

        logger.info(f"Project root detected: {project_root}")
        logger.info(f"Migrating all CSV files in: {data_dir}")
        logger.info(f"Full path: {data_dir.absolute()}")
        logger.info("=" * 50)

        if not data_dir.exists():
            logger.error(f"Directory does not exist: {data_dir.absolute()}")
            logger.error("Please ensure you're running from the correct location")
            sys.exit(1)

        stats = migrator.migrate_directory(data_dir)
    elif args.directory:
        # Migrate directory
        stats = migrator.migrate_directory(Path(args.directory), args.recursive)
    else:
        # Migrate single file
        migrator.stats['total'] = 1
        if migrator.migrate_file(Path(args.csv_file)):
            migrator.stats['success'] = 1
        else:
            migrator.stats['failed'] = 1

    # Print summary
    migrator.print_summary()


if __name__ == "__main__":
    main()